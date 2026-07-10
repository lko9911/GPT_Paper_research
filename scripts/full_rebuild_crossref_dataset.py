"""Update the paper dataset from Crossref search results.

The updater preserves the existing active paper set because many records contain
paid OpenAI summaries. New Crossref results are merged by DOI/title key; existing
papers that are not returned by the latest Crossref search remain active.
OpenAlex is used only for DOI-based corresponding-author completion when
Crossref does not provide a corresponding-author flag.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import sys
import time
import zipfile
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from fetch_crossref import fetch_crossref, fetch_crossref_by_issn_query
from fetch_openalex import fetch_openalex_by_doi
from summarize import summarize_record
from update_papers import (
    CURATED_MIN_SCORE,
    _is_plausible,
    _journal_quality,
    _venue_classification,
    _safe_year,
    _split_curated_archive,
    _strip_transient,
)

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
PAPERS_PATH = DATA_DIR / "papers.json"
ARCHIVE_PAPERS_PATH = DATA_DIR / "archive_papers.json"
PAPERS_CSV_PATH = DATA_DIR / "papers.csv"
PAPERS_XLSX_PATH = DATA_DIR / "papers.xlsx"
SITE_META_PATH = DATA_DIR / "site_meta.json"
QUERIES_PATH = DATA_DIR / "queries.json"
CROSSREF_VENUE_QUERIES_PATH = DATA_DIR / "crossref_venue_queries.json"
BACKUP_ROOT = DATA_DIR / "old_exports"
DEFAULT_SINCE_YEAR = 2024
SEARCH_ROWS = int(os.getenv("SEARCH_PER_PAGE", "200"))
VENUE_SEARCH_ROWS = int(os.getenv("CROSSREF_VENUE_SEARCH_ROWS", str(SEARCH_ROWS)))


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    today = date.today().isoformat()
    run_started_at = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    since_year = _since_year()

    print("Incremental merge mode enabled")
    print("Existing active paper dataset will be preserved")
    print("Priority venue search disabled")
    print("OpenAlex general search disabled")
    existing_active = _load_json(PAPERS_PATH, [])
    existing_archive = _load_json(ARCHIVE_PAPERS_PATH, [])
    existing_lineage = _existing_date_lineage()
    existing_openai_summaries = _existing_openai_summaries()
    backup_dir = _archive_existing_outputs(run_started_at)
    print(f"Existing paper dataset archived: {backup_dir.relative_to(ROOT)}")
    print(f"Existing active papers loaded for preservation: {len(existing_active)}")

    queries = _filtered_queries(_load_json(QUERIES_PATH, []))
    venue_targets = _load_json(CROSSREF_VENUE_QUERIES_PATH, [])
    print("Crossref-only search started")
    raw_candidates = _collect_crossref_candidates(queries, since_year)
    venue_candidates = _collect_crossref_venue_candidates(venue_targets, since_year)
    raw_candidates.extend(venue_candidates)
    print(f"Crossref-only search completed: {len(raw_candidates)} raw records")

    deduped = _dedupe_crossref_records(raw_candidates)
    print(f"DOI/title/year/first-author de-duplication complete: {len(deduped)} records")

    papers = []
    openalex_stats = {
        "checked": 0,
        "completed": 0,
        "authors_completed": 0,
        "not_found": 0,
        "skipped_has_crossref_corresponding": 0,
        "skipped_no_doi": 0,
    }
    print("OpenAlex used only for missing corresponding author cross-check")
    for record in deduped:
        summarized = summarize_record(record, allow_openai=False)
        completed = _complete_corresponding_author_from_openalex(summarized, openalex_stats)
        papers.append(_finalize_crossref_record(completed, today, existing_lineage, existing_openai_summaries))

    new_cleaned = [_strip_transient(paper) for paper in papers]
    merged_input, merge_stats = _merge_existing_active_with_new(existing_active, new_cleaned, today)
    cleaned = [_strip_transient(paper) for paper in merged_input]
    curated, new_archive, split_stats = _split_curated_archive(cleaned)
    archive = _merge_archives(existing_archive, new_archive, curated)
    weekly_added = sum(1 for paper in curated if paper.get("is_weekly_new"))
    _write_json(PAPERS_PATH, curated)
    _write_json(ARCHIVE_PAPERS_PATH, archive)
    _export_csv(PAPERS_CSV_PATH, curated)
    _export_xlsx(PAPERS_XLSX_PATH, curated)

    _write_json(
        SITE_META_PATH,
        {
            "last_run_at_utc": run_started_at,
            "last_run_date": today,
            "paper_count": len(curated),
            "curated_count": len(curated),
            "raw_candidate_count": len(new_cleaned),
            "merged_candidate_count": len(cleaned),
            "archived_count": len(archive),
            "new_archived_count": len(new_archive),
            "hidden_low_relevance_count": split_stats["low_relevance"],
            "hidden_low_venue_trust_count": split_stats.get("low_venue_trust", 0),
            "duplicate_archived_count": split_stats["duplicate_title"],
            "papers_added": merge_stats["new_added"],
            "raw_records_added": len(new_cleaned),
            "weekly_added_count": weekly_added,
            "weekly_window_days": 7,
            "since_year": since_year,
            "curated_min_score": CURATED_MIN_SCORE,
            "sources": ["Crossref"],
            "collection_mode": "incremental_crossref_merge",
            "existing_active_preserved": True,
            "existing_active_count_before": len(existing_active),
            "new_crossref_curated_candidates": len(new_cleaned),
            "merge_stats": merge_stats,
            "openalex_general_search_enabled": False,
            "priority_venue_search_enabled": False,
            "openalex_used_for": "missing_corresponding_author_doi_cross_check_only",
            "openalex_corresponding_author_stats": openalex_stats,
            "backup_dir": str(backup_dir.relative_to(ROOT)),
        },
    )

    print("Incremental Crossref-based dataset exported")
    print(
        "Incremental update complete. "
        f"Curated {len(curated)} papers; added {merge_stats['new_added']} new papers; "
        f"updated {merge_stats['existing_updated']} existing papers; preserved "
        f"{merge_stats['existing_preserved_without_current_match']} existing papers without a current Crossref hit. "
        f"OpenAlex checked {openalex_stats['checked']} DOI records and completed "
        f"{openalex_stats['completed']} corresponding-author entries."
    )


def _collect_crossref_candidates(queries: list[str], since_year: int) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for query in queries:
        print(f"Crossref search query: {query}")
        try:
            records = fetch_crossref(query, rows=SEARCH_ROWS, from_year=since_year)
        except Exception as exc:
            print(f"Crossref search failed for {query!r}: {exc}")
            continue
        for record in records:
            record["source"] = ["Crossref"]
            record["metadata_source"] = "crossref"
            record["openalex_checked"] = False
            record["openalex_used_for"] = None
            record["corresponding_author_source"] = "crossref" if record.get("corresponding_authors") else None
            if _is_plausible(record, since_year):
                candidates.append(record)
        time.sleep(float(os.getenv("API_SLEEP_SECONDS", "0.2")))
    return candidates


def _merge_existing_active_with_new(
    existing_active: list[dict[str, Any]],
    new_records: list[dict[str, Any]],
    today: str,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Keep existing active papers and merge in newly observed Crossref papers.

    The previous full-rebuild behavior allowed a paper to disappear from the
    public site when Crossref relevance results changed. That is not acceptable
    for paid/manual OpenAI summaries, so active records are now sticky unless a
    later curation rule explicitly archives them.
    """

    existing_by_key: dict[str, dict[str, Any]] = {}
    for record in existing_active:
        key = _dedupe_key(record)
        if key:
            existing_by_key[key] = _refresh_preserved_existing_record(record, today)

    merged_by_key = dict(existing_by_key)
    stats = {
        "existing_active_before": len(existing_by_key),
        "new_crossref_records": len(new_records),
        "existing_updated": 0,
        "new_added": 0,
        "existing_preserved_without_current_match": 0,
    }
    seen_current_keys: set[str] = set()

    for record in new_records:
        key = _dedupe_key(record)
        if not key:
            continue
        seen_current_keys.add(key)
        existing = existing_by_key.get(key)
        if existing:
            merged_by_key[key] = _merge_existing_and_current_record(existing, record, today)
            stats["existing_updated"] += 1
        else:
            current = dict(record)
            current["not_seen_in_latest_crossref_run"] = False
            current["last_seen_in_crossref_run"] = today
            merged_by_key[key] = current
            stats["new_added"] += 1

    missing_current = set(existing_by_key) - seen_current_keys
    stats["existing_preserved_without_current_match"] = len(missing_current)
    for key in missing_current:
        merged_by_key[key]["not_seen_in_latest_crossref_run"] = True

    return list(merged_by_key.values()), stats


def _refresh_preserved_existing_record(record: dict[str, Any], today: str) -> dict[str, Any]:
    refreshed = dict(record)
    refreshed["is_weekly_new"] = _is_date_within_days(str(refreshed.get("first_added") or ""), today, 7)
    refreshed.setdefault("last_seen_in_crossref_run", refreshed.get("last_updated", ""))
    refreshed.setdefault("not_seen_in_latest_crossref_run", False)
    return refreshed


def _merge_existing_and_current_record(
    existing: dict[str, Any],
    current: dict[str, Any],
    today: str,
) -> dict[str, Any]:
    merged = dict(existing)
    merged.update(current)
    merged["first_added"] = existing.get("first_added") or current.get("first_added") or today
    merged["last_updated"] = today
    merged["is_weekly_new"] = _is_date_within_days(str(merged.get("first_added") or ""), today, 7)
    merged["last_seen_in_crossref_run"] = today
    merged["not_seen_in_latest_crossref_run"] = False

    if existing.get("summary_provider") == "openai" or existing.get("openai_summary_applied") is True:
        for key in (
            "ai_summary_en",
            "relevance_note_en",
            "categories",
            "tags",
            "relevance_score",
            "summary_provider",
            "openai_summary_applied",
            "abstract_used_for_summary",
        ):
            if existing.get(key) not in (None, "", []):
                merged[key] = existing[key]
        merged["summary_provider"] = "openai"
        merged["openai_summary_applied"] = True

    return merged


def _merge_archives(
    existing_archive: list[dict[str, Any]],
    new_archive: list[dict[str, Any]],
    curated: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    active_keys = {_dedupe_key(record) for record in curated if _dedupe_key(record)}
    merged: dict[str, dict[str, Any]] = {}
    for record in existing_archive + new_archive:
        key = _dedupe_key(record)
        if not key or key in active_keys:
            continue
        if key not in merged:
            merged[key] = record
        else:
            merged[key] = _better_archive_record(merged[key], record)
    return sorted(merged.values(), key=_archive_sort_key, reverse=True)


def _better_archive_record(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
    def score(record: dict[str, Any]) -> tuple[int, int, int, int]:
        return (
            1 if record.get("summary_provider") == "openai" or record.get("openai_summary_applied") else 0,
            1 if record.get("doi") else 0,
            int(record.get("relevance_score") or 0),
            int(record.get("year") or 0),
        )

    return right if score(right) > score(left) else left


def _archive_sort_key(record: dict[str, Any]) -> tuple[int, int, str]:
    return (
        int(record.get("year") or 0),
        int(record.get("relevance_score") or 0),
        str(record.get("title") or ""),
    )


def _collect_crossref_venue_candidates(venue_targets: list[dict[str, Any]], since_year: int) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    if not venue_targets:
        print("Crossref ISSN venue search disabled: no venue query file entries")
        return candidates
    if _env_flag("SKIP_CROSSREF_VENUE_QUERIES"):
        print("Crossref ISSN venue search skipped because SKIP_CROSSREF_VENUE_QUERIES is enabled")
        return candidates

    print("Crossref ISSN venue search started")
    for target in venue_targets:
        venue_name = str(target.get("name") or "Unknown venue")
        issns = [str(issn).strip() for issn in target.get("issn", []) if str(issn).strip()]
        queries = [str(query).strip() for query in target.get("queries", []) if str(query).strip()]
        if not issns or not queries:
            print(f"Crossref ISSN venue search skipped for {venue_name}: missing ISSN or queries")
            continue
        for issn in issns:
            for query in queries:
                print(f"Crossref venue search: {venue_name} / ISSN {issn} / {query}")
                try:
                    records = fetch_crossref_by_issn_query(
                        query,
                        issn=issn,
                        rows=VENUE_SEARCH_ROWS,
                        from_year=since_year,
                    )
                except Exception as exc:
                    print(f"Crossref venue search failed for {venue_name} / {issn} / {query!r}: {exc}")
                    continue
                for record in records:
                    record["source"] = ["Crossref"]
                    record["metadata_source"] = "crossref"
                    record["crossref_collection_route"] = "issn_venue_query"
                    record["crossref_target_venue"] = venue_name
                    record["crossref_target_issn"] = issn
                    record["openalex_checked"] = False
                    record["openalex_used_for"] = None
                    record["corresponding_author_source"] = "crossref" if record.get("corresponding_authors") else None
                    if _is_plausible(record, since_year):
                        candidates.append(record)
                time.sleep(float(os.getenv("API_SLEEP_SECONDS", "0.2")))
    print(f"Crossref ISSN venue search completed: {len(candidates)} plausible records")
    return candidates


def _dedupe_crossref_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for record in records:
        key = _dedupe_key(record)
        if key not in index:
            index[key] = record
            continue
        index[key] = _better_record(index[key], record)
    return list(index.values())


def _dedupe_key(record: dict[str, Any]) -> str:
    doi = _clean_doi(record.get("doi"))
    if doi:
        return f"doi:{doi}"
    title = _normalize_title(record.get("title", ""))
    year = str(record.get("year") or "")
    first_author = _normalize_title((record.get("authors") or [""])[0])
    return f"title-year-author:{title}|{year}|{first_author}"


def _existing_date_lineage() -> dict[str, str]:
    lineage: dict[str, str] = {}
    for record in _load_json(PAPERS_PATH, []) + _load_json(ARCHIVE_PAPERS_PATH, []):
        key = _dedupe_key(record)
        first_added = str(record.get("first_added") or "").strip()
        if key and first_added:
            lineage[key] = first_added
    return lineage


def _existing_openai_summaries() -> dict[str, dict[str, Any]]:
    summaries: dict[str, dict[str, Any]] = {}
    for record in _load_json(PAPERS_PATH, []) + _load_json(ARCHIVE_PAPERS_PATH, []):
        if record.get("summary_provider") != "openai" and record.get("openai_summary_applied") is not True:
            continue
        key = _dedupe_key(record)
        summary = str(record.get("ai_summary_en") or "").strip()
        if not key or not summary:
            continue
        summaries[key] = {
            "ai_summary_en": summary,
            "relevance_note_en": record.get("relevance_note_en", ""),
            "tags": record.get("tags", []),
            "categories": record.get("categories", []),
            "relevance_score": record.get("relevance_score"),
            "summary_provider": "openai",
            "openai_summary_applied": True,
            "abstract_used_for_summary": bool(record.get("abstract_used_for_summary")),
        }
    print(f"Existing OpenAI summaries available for preservation: {len(summaries)}")
    return summaries


def _better_record(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
    def score(record: dict[str, Any]) -> tuple[int, int, int, int, int]:
        return (
            1 if record.get("doi") else 0,
            1 if record.get("_abstract") else 0,
            1 if record.get("corresponding_authors") else 0,
            len(record.get("author_details") or []),
            len(record.get("authors") or []),
        )

    winner = right if score(right) > score(left) else left
    loser = left if winner is right else right
    for key in ("_abstract", "author_details", "corresponding_authors", "authors", "issn", "publisher"):
        if not winner.get(key) and loser.get(key):
            winner[key] = loser[key]
    return winner


def _complete_corresponding_author_from_openalex(record: dict[str, Any], stats: dict[str, int]) -> dict[str, Any]:
    if record.get("corresponding_authors"):
        stats["skipped_has_crossref_corresponding"] += 1
        record["corresponding_author_source"] = "crossref"
        record["openalex_checked"] = False
        record["openalex_used_for"] = None
        return record

    doi = _clean_doi(record.get("doi"))
    if not doi:
        stats["skipped_no_doi"] += 1
        record["corresponding_author_source"] = None
        record["openalex_checked"] = False
        record["openalex_used_for"] = None
        return record

    stats["checked"] += 1
    print(f"OpenAlex checked by DOI for missing corresponding author: {doi}")
    try:
        openalex_record = fetch_openalex_by_doi(doi)
    except Exception as exc:
        stats["not_found"] += 1
        print(f"OpenAlex DOI cross-check failed for {doi}: {exc}")
        record["openalex_checked"] = True
        record["openalex_used_for"] = None
        record["corresponding_author_source"] = None
        return record

    record["openalex_checked"] = True
    if not openalex_record or _clean_doi(openalex_record.get("doi")) != doi:
        stats["not_found"] += 1
        record["openalex_used_for"] = None
        record["corresponding_author_source"] = None
        print(f"corresponding author not found in Crossref or OpenAlex: {doi}")
        return record

    record["openalex_crosscheck_work_id"] = openalex_record.get("openalex_work_id", "")
    _merge_openalex_author_metadata(record, openalex_record, stats)
    corresponding = openalex_record.get("corresponding_authors") or []
    if corresponding:
        record["corresponding_authors"] = corresponding
        record["corresponding_author_available"] = True
        record["corresponding_author_source"] = "openalex_cross_check"
        record["openalex_used_for"] = "corresponding_author_completion"
        stats["completed"] += 1
        print(f"corresponding author completed from OpenAlex: {doi}")
        return record

    stats["not_found"] += 1
    record["openalex_used_for"] = None
    record["corresponding_author_source"] = None
    print(f"corresponding author not found in Crossref or OpenAlex: {doi}")
    return record


def _merge_openalex_author_metadata(
    record: dict[str, Any],
    openalex_record: dict[str, Any],
    stats: dict[str, int],
) -> None:
    copied = False
    if openalex_record.get("authors") and not record.get("authors"):
        record["authors"] = openalex_record["authors"]
        copied = True
    if openalex_record.get("author_details") and not record.get("author_details"):
        record["author_details"] = openalex_record["author_details"]
        copied = True
    if copied:
        stats["authors_completed"] = stats.get("authors_completed", 0) + 1


def _finalize_crossref_record(
    record: dict[str, Any],
    today: str,
    existing_lineage: dict[str, str],
    existing_openai_summaries: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    doi = _clean_doi(record.get("doi"))
    title = record.get("title", "Untitled")
    paper_id = doi or _title_hash(title)
    dedupe_key = _dedupe_key(record)
    first_added = existing_lineage.get(dedupe_key, today)
    preserved_summary = existing_openai_summaries.get(dedupe_key, {})
    year = _safe_year(record.get("year"))
    journal_quality = _journal_quality(record)
    is_core = _is_manual_core_journal(journal_quality)
    venue_classification = _venue_classification(record)
    finalized = {
        "id": paper_id,
        "title": title,
        "authors": record.get("authors", []),
        "author_details": record.get("author_details", []),
        "corresponding_authors": record.get("corresponding_authors", []),
        "corresponding_author_available": bool(record.get("corresponding_authors")),
        "corresponding_author_source": record.get("corresponding_author_source"),
        "year": year,
        "venue": record.get("venue", ""),
        "doi": doi,
        "url": record.get("url") or (f"https://doi.org/{doi}" if doi else ""),
        "source": ["Crossref"],
        "metadata_source": "crossref",
        "crossref_type": record.get("crossref_type", ""),
        "crossref_collection_route": record.get("crossref_collection_route", "keyword_query"),
        "crossref_target_venue": record.get("crossref_target_venue", ""),
        "crossref_target_issn": record.get("crossref_target_issn", ""),
        "issn": record.get("issn", []),
        "issn_l": record.get("issn_l", ""),
        "publisher": record.get("publisher", ""),
        "openalex_checked": bool(record.get("openalex_checked")),
        "openalex_used_for": record.get("openalex_used_for"),
        "openalex_crosscheck_work_id": record.get("openalex_crosscheck_work_id", ""),
        "openalex_work_id": "",
        "openalex_source_id": "",
        "venue_metrics": {},
        "journal_quality": journal_quality,
        "publication_type": venue_classification["publication_type"],
        "venue_trust": venue_classification["venue_trust"],
        "venue_trust_reason": venue_classification["venue_trust_reason"],
        "is_core_venue": is_core,
        "core_status": "core" if is_core else "non-core",
        "venue_scope": "core" if is_core else "non-core",
        "core_source": "manual_core_venue" if is_core else "schema_placeholder_crossref_rebuild",
        "categories": (preserved_summary.get("categories") or record.get("categories", ["Multi-material AM"]))[:2],
        "tags": (preserved_summary.get("tags") or record.get("tags", []))[:6],
        "relevance_score": int(preserved_summary.get("relevance_score") or record.get("relevance_score", 5)),
        "curation_priority": False,
        "ai_summary_en": preserved_summary.get("ai_summary_en") or record.get("ai_summary_en", ""),
        "summary_provider": preserved_summary.get("summary_provider") or record.get("_summary_provider", "fallback"),
        "openai_summary_applied": bool(preserved_summary.get("openai_summary_applied")),
        "relevance_note_en": preserved_summary.get("relevance_note_en") or record.get("relevance_note_en", ""),
        "abstract_used_for_summary": bool(preserved_summary.get("abstract_used_for_summary") or record.get("_abstract")),
        "raw_abstract_displayed": False,
        "pdf_stored": False,
        "first_added": first_added,
        "last_updated": today,
        "is_weekly_new": _is_date_within_days(first_added, today, 7),
    }
    finalized["journal_quality"] = _journal_quality(finalized)
    venue_classification = _venue_classification(finalized)
    finalized["publication_type"] = venue_classification["publication_type"]
    finalized["venue_trust"] = venue_classification["venue_trust"]
    finalized["venue_trust_reason"] = venue_classification["venue_trust_reason"]
    is_core = _is_manual_core_journal(finalized["journal_quality"])
    finalized["is_core_venue"] = is_core
    finalized["core_status"] = "core" if is_core else "non-core"
    finalized["venue_scope"] = "core" if is_core else "non-core"
    finalized["core_source"] = "manual_core_venue" if is_core else "schema_placeholder_crossref_rebuild"
    return finalized


def _is_date_within_days(value: str, reference: str, days: int) -> bool:
    try:
        reference_date = date.fromisoformat(reference[:10])
        value_date = date.fromisoformat(str(value or "")[:10])
    except ValueError:
        return False
    start = reference_date.toordinal() - max(0, days - 1)
    return start <= value_date.toordinal() <= reference_date.toordinal()


def _is_manual_core_journal(journal_quality: dict[str, Any]) -> bool:
    return journal_quality.get("basis") == "manual_core_venue"


def _export_csv(path: Path, records: list[dict[str, Any]]) -> None:
    fieldnames = _export_fieldnames()
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for record in records:
            writer.writerow({key: _csv_value(record.get(key)) for key in fieldnames})
    print(f"Wrote {path.relative_to(ROOT)}")


def _export_xlsx(path: Path, records: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        print(f"Skipping XLSX export because openpyxl is not installed: {exc}")
        return

    fieldnames = _export_fieldnames()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "papers"
    sheet.append(fieldnames)
    for record in records:
        sheet.append([_csv_value(record.get(key)) for key in fieldnames])
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    widths = {
        "A": 22,
        "B": 56,
        "C": 42,
        "D": 10,
        "E": 34,
        "F": 28,
        "G": 36,
        "H": 18,
        "I": 18,
    }
    for index in range(1, len(fieldnames) + 1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = widths.get(letter, 18)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    workbook.save(path)
    print(f"Wrote {path.relative_to(ROOT)}")


def _export_fieldnames() -> list[str]:
    return [
        "id",
        "title",
        "authors",
        "year",
        "venue",
        "doi",
        "url",
        "source",
        "metadata_source",
        "crossref_type",
        "crossref_collection_route",
        "crossref_target_venue",
        "crossref_target_issn",
        "issn",
        "issn_l",
        "publisher",
        "publication_type",
        "venue_trust",
        "venue_trust_reason",
        "openalex_source_id",
        "openalex_venue_rank",
        "openalex_venue_rank_number",
        "openalex_venue_rank_score",
        "openalex_venue_rank_percentile",
        "openalex_venue_rank_basis",
        "corresponding_authors",
        "corresponding_author_available",
        "corresponding_author_source",
        "openalex_checked",
        "openalex_used_for",
        "openalex_crosscheck_work_id",
        "is_core_venue",
        "core_status",
        "venue_scope",
        "core_source",
        "categories",
        "tags",
        "relevance_score",
        "summary_provider",
        "openai_summary_applied",
        "abstract_used_for_summary",
        "raw_abstract_displayed",
        "pdf_stored",
        "first_added",
        "last_updated",
    ]


def _csv_value(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _archive_existing_outputs(run_started_at: str) -> Path:
    stamp = re.sub(r"[^0-9]", "", run_started_at)[:14]
    backup_dir = BACKUP_ROOT / f"full_rebuild_{stamp}"
    backup_dir.mkdir(parents=True, exist_ok=True)
    targets = [
        PAPERS_PATH,
        ARCHIVE_PAPERS_PATH,
        PAPERS_CSV_PATH,
        PAPERS_XLSX_PATH,
        DATA_DIR / "papers_index.json",
        DATA_DIR / "archive_papers_index.json",
        DATA_DIR / "detail_manifest.json",
        DATA_DIR / "archive_detail_manifest.json",
        SITE_META_PATH,
        DATA_DIR / "details",
        DATA_DIR / "archive_details",
    ]
    manifest = []
    for target in targets:
        if not target.exists():
            continue
        manifest.append(str(target.relative_to(ROOT)))
        if target.is_dir():
            _zip_dir(target, backup_dir / f"{target.name}.zip")
        else:
            _zip_file(target, backup_dir / f"{target.name}.zip")
    (backup_dir / "MANIFEST.json").write_text(
        json.dumps(
            {
                "created_at_utc": run_started_at,
                "mode": "incremental_crossref_merge",
                "archived_paths": manifest,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    return backup_dir


def _zip_file(path: Path, output: Path) -> None:
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.write(path, arcname=path.name)


def _zip_dir(path: Path, output: Path) -> None:
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for child in sorted(path.rglob("*")):
            if child.is_file():
                archive.write(child, arcname=str(child.relative_to(path.parent)))


def _filtered_queries(queries: list[str]) -> list[str]:
    query_filter = os.getenv("UPDATE_QUERY_FILTER", "").strip()
    if not query_filter:
        return queries
    needles = [part.strip().lower() for part in query_filter.split(",") if part.strip()]
    filtered = [query for query in queries if any(needle in query.lower() for needle in needles)]
    print(f"UPDATE_QUERY_FILTER selected {len(filtered)} of {len(queries)} queries.")
    return filtered


def _since_year() -> int:
    value = os.getenv("SINCE_YEAR", str(DEFAULT_SINCE_YEAR))
    try:
        return int(value)
    except ValueError:
        print(f"Invalid SINCE_YEAR={value!r}; using {DEFAULT_SINCE_YEAR}")
        return DEFAULT_SINCE_YEAR


def _env_flag(name: str) -> bool:
    return os.getenv(name, "").strip().lower() in {"1", "true", "yes", "on"}


def _load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {path.relative_to(ROOT)}")


def _clean_doi(value: Any) -> str:
    if not value:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"^https?://(dx\.)?doi\.org/", "", text)
    return re.sub(r"(\.pdf|/pdf)$", "", text).strip()


def _normalize_title(title: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(title or "").lower()).strip()


def _title_hash(title: str) -> str:
    return hashlib.sha256(_normalize_title(title).encode("utf-8")).hexdigest()[:16]


if __name__ == "__main__":
    main()
